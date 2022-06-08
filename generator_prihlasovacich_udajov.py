from random import choice, randrange
from openpyxl import load_workbook
wb = load_workbook(filename='D:\\ucty.xlsx')
ws = wb['Hárok1']
import secrets
import string
numbers = '0123456789'


def spravna_spoluhlaska():
    pravidelnost1spolu = 'tnshrdl'
    pravidelnost2spolu = 'cmfgpwbvk'
    pravidelnost3spolu = 'xjqz'
    spravne_cislo = randrange(1, 13)
    if spravne_cislo % 5 == 0:
        return choice(pravidelnost2spolu)
    elif spravne_cislo % 9 == 0:
        if randrange(1, 3) % 2 == 0:
            return choice(pravidelnost3spolu)
        else:
            return choice(pravidelnost1spolu)
    else:
        return choice(pravidelnost1spolu)


def spravna_samohlaska():
    pravidelnost1samo = 'ea'
    pravidelnost2samo = 'oi'
    pravidelnost3samo = 'uy'
    spravne_cislo = randrange(1, 13)
    if spravne_cislo % 5 == 0:
        return choice(pravidelnost2samo)
    elif spravne_cislo % 9 == 0:
        if randrange(1, 3) % 2 == 0:
            return choice(pravidelnost3samo)
        else:
            return choice(pravidelnost1samo)
    else:
        return choice(pravidelnost1samo)


def name(dlzka, pridat_cisla=True):
    name = ''
    pismealebocislo = round(randrange(round(dlzka * 0.6), round(dlzka * 0.9)))
    kroky = 0
    while len(name) <= dlzka:
        if randrange(5) % 5 == 0:
            kroky = 1
        while len(name) <= pismealebocislo:
            if (kroky % 3) == 0:
                name += spravna_spoluhlaska()
            elif (kroky % 3) == 1:
                name += spravna_samohlaska()
            elif (kroky % 3) == 2:
                if randrange(3) % 3 == 0:
                    name += spravna_spoluhlaska()
                else:
                    name += ''
            kroky += 1
        name += choice(numbers)
    if not pridat_cisla:
        name = ''.join([i for i in name if not i.isdigit()])
    return name

def heslo(dlzka):
    alphabet = string.ascii_letters + string.digits + string.punctuation
    password = ''.join(secrets.choice(alphabet) for i in range(dlzka))
    if password[0]!="=" and password!=None:
        return password
    else:
        heslo(dlzka)

def oprava_hesiel(dlzka_hesla):
    row = 5
    while row <= ws.max_row:
        if ws['C' + str(row)].value == None:
            ws.cell(column=3, row=row, value=heslo(dlzka_hesla))

        if ws['E' + str(row)].value == None:
            ws.cell(column=5, row=row, value=heslo(dlzka_hesla))
        row += 1

def zapis_ucty(pocet_uctov, dlzka_hesla, dlzka_uctu):
    if ws.max_row==4:
        meno=input()
        if string.digits in meno:
            skusobne_meno = name(dlzka_uctu, False)
            print(skusobne_meno)
            while int(input("")) != 1:
                skusobne_meno = name(dlzka_uctu, False)
                print(skusobne_meno)
            meno = (skusobne_meno + 'boost')
        else:
            meno+="boost"
    else:
        meno = ws['B' + str(ws.max_row)].value
        meno = meno[:(meno.rfind("boost"))]
        meno+="boost"

    newrowlocation = ws.max_row + 1
    for i in range(pocet_uctov):
        ws.cell(column=1, row=newrowlocation, value=str(newrowlocation - 5))
        ws.cell(column=2, row=newrowlocation, value=meno + str(newrowlocation - 5))
        ws.cell(column=3, row=newrowlocation, value=heslo(dlzka_hesla))
        ws.cell(column=4, row=newrowlocation, value=(meno + str(newrowlocation - 5) + '@gmail.com'))
        ws.cell(column=5, row=newrowlocation, value=heslo(dlzka_hesla))
        newrowlocation = ws.max_row + 1

    for i in range(5):
        oprava_hesiel(dlzka_hesla)

def vymaz_vsetky_ucty():
    ws.delete_rows(6, ws.max_row+1)

#zapis_ucty(15,16,6)
for i in range(100):
    print(name(7,pridat_cisla=False))

wb.save(filename='D:\\ucty.xlsx')
wb.close()
